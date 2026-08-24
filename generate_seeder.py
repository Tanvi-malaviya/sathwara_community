import openpyxl
import re
import json

wb = openpyxl.load_workbook('SS - Website Data.xlsx')
ws = wb['newmemberapproved']

records = []
seen_member_codes = {}
seen_emails = set()

for r in range(2, ws.max_row + 1):
    member_code = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ''
    surname = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else ''
    first_name = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else ''
    middle_name = str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else ''
    email = str(ws.cell(r, 6).value).strip() if ws.cell(r, 6).value else ''
    phone = str(ws.cell(r, 7).value).strip() if ws.cell(r, 7).value else ''
    address = str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else ''
    area = str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else ''

    if not member_code and not first_name and not surname and not email:
        continue

    # Clean phone (remove decimals if float, spaces)
    if phone.endswith('.0'):
        phone = phone[:-2]
    phone = re.sub(r'[^\d]', '', phone)

    # Clean Area (strip trailing pincode like " - 380024" or " 380024")
    clean_area = area
    if clean_area:
        clean_area = re.sub(r'\s*-\s*\d{6}\s*$', '', clean_area).strip()
        clean_area = re.sub(r'\s+\d{6}\s*$', '', clean_area).strip()

    # Handle duplicate member codes if any
    if member_code in seen_member_codes:
        seen_member_codes[member_code] += 1
        member_code = f"{member_code}_{seen_member_codes[member_code]}"
    else:
        seen_member_codes[member_code] = 1

    # Clean / validate email
    clean_email = email.lower().strip() if email else ''
    if not clean_email or '@' not in clean_email or '.' not in clean_email.split('@')[-1]:
        clean_email = ''
    elif clean_email in seen_emails:
        clean_email = ''
    else:
        seen_emails.add(clean_email)

    records.append({
        'member_code': member_code,
        'surname': surname,
        'first_name': first_name,
        'middle_name': middle_name,
        'email': clean_email,
        'phone': phone,
        'address': address,
        'area': clean_area
    })

print(f"Total parsed records: {len(records)}")

def var_export_php(arr):
    lines = ["[\n"]
    for item in arr:
        lines.append("    [\n")
        for k, v in item.items():
            safe_v = (v or "").replace("\\", "\\\\").replace("'", "\\'")
            lines.append(f"        '{k}' => '{safe_v}',\n")
        lines.append("    ],\n")
    lines.append("]")
    return "".join(lines)

records_export = var_export_php(records)

php_template = r"""<?php

namespace Database\Seeders;

use Illuminate\Database\Seeder;
use App\Models\User;
use App\Models\MemberProfile;
use App\Models\Area;
use App\Models\Business;
use Spatie\Permission\Models\Role;
use Illuminate\Support\Facades\Hash;
use Illuminate\Support\Facades\DB;

class MemberImportSeeder extends Seeder
{
    /**
     * Run database seeds for Member dataset.
     */
    public function run(): void
    {
        $memberRole = Role::firstOrCreate(['name' => 'Member']);
        $defaultPassword = Hash::make('password');

        $records = """ + records_export + r""";

        // Canonical area aliases mapping for deduplication
        $areaAliases = [
            'naroda road'   => 'Naroda',
            'naroda'        => 'Naroda',
            'new naroda'    => 'Nava Naroda',
            'nava naroda'   => 'Nava Naroda',
            'navanaroda'    => 'Nava Naroda',
            'newnaroda'     => 'Nava Naroda',
            'taltej'        => 'Thaltej',
            'thaltej'       => 'Thaltej',
            'vasana'        => 'Vasna',
            'vasna'         => 'Vasna',
            'niranaynagar'  => 'Nirnaynagar',
            'nirnaynagar'   => 'Nirnaynagar',
            'sahibag'       => 'Shahibaug',
            'shahibaug'     => 'Shahibaug',
            'satelite'      => 'Satellite',
            'satellite'     => 'Satellite',
            'wadj'          => 'Vadaj',
            'vadaj'         => 'Vadaj',
        ];

        // Canonical pincodes
        $canonicalPincodes = [
            'naroda'       => '382330',
            'nava naroda'  => '382330',
            'thaltej'      => '380059',
            'vasna'        => '380007',
            'nirnaynagar'  => '382481',
            'shahibaug'    => '380004',
            'satellite'    => '380015',
            'bapunagar'    => '380024',
            'nikol'        => '382350',
            'vastral'      => '382418',
            'ranip'        => '382480',
            'chandlodia'   => '382481',
        ];

        DB::beginTransaction();

        try {
            // 1. Ensure canonical primary areas exist
            $canonicalAreaModels = [];
            foreach (array_unique(array_values($areaAliases)) as $canonicalName) {
                $cKey = strtolower($canonicalName);
                $existing = Area::whereRaw('LOWER(name) = ?', [$cKey])->first();
                $pin = $canonicalPincodes[$cKey] ?? null;
                if (!$existing) {
                    $existing = Area::create([
                        'name'    => $canonicalName,
                        'pincode' => $pin,
                    ]);
                } elseif ($pin && empty($existing->pincode)) {
                    $existing->update(['pincode' => $pin]);
                }
                $canonicalAreaModels[$cKey] = $existing;
            }

            // 2. Merge duplicate areas and delete duplicate entries
            foreach ($areaAliases as $aliasName => $targetCanonicalName) {
                if (strtolower($aliasName) === strtolower($targetCanonicalName)) {
                    continue;
                }
                $targetArea = $canonicalAreaModels[strtolower($targetCanonicalName)] ?? null;
                if (!$targetArea) continue;

                $dupAreas = Area::whereRaw('LOWER(name) = ?', [strtolower($aliasName)])
                    ->where('id', '!=', $targetArea->id)
                    ->get();

                foreach ($dupAreas as $da) {
                    MemberProfile::where('area_id', $da->id)->update(['area_id' => $targetArea->id]);
                    Business::where('area_id', $da->id)->update(['area_id' => $targetArea->id]);
                    $da->delete();
                }
            }

            // 3. Preload all active areas
            $allAreas = Area::all();
            $areaMap = [];
            foreach ($allAreas as $a) {
                $areaMap[strtolower(trim($a->name))] = $a;
            }

            $count = 0;
            $importedUserIds = [];

            foreach ($records as $item) {
                $memberCode = trim($item['member_code']);
                $surname = trim($item['surname']);
                $firstName = trim($item['first_name']);
                $middleName = trim($item['middle_name']);
                $email = !empty($item['email']) ? strtolower(trim($item['email'])) : null;
                $phone = trim($item['phone']);
                $rawArea = trim($item['area']);
                $address = trim($item['address']);

                // Full Name
                $nameParts = array_filter([$surname, $firstName, $middleName]);
                $fullName = implode(' ', $nameParts);
                if (empty($fullName)) {
                    $fullName = $memberCode ?: 'Community Member';
                }

                // Resolve Area Name through Alias Mapping
                $resolvedAreaName = $rawArea;
                $lowRaw = strtolower($rawArea);
                if (isset($areaAliases[$lowRaw])) {
                    $resolvedAreaName = $areaAliases[$lowRaw];
                }

                // Specific address-level distinction for Naroda vs Nava Naroda
                $lowAddr = strtolower($address);
                if (str_contains($lowAddr, 'new naroda') || str_contains($lowAddr, 'nava naroda') || str_contains($lowAddr, 'નવા નરોડા')) {
                    $resolvedAreaName = 'Nava Naroda';
                } elseif (str_contains($lowAddr, 'naroda road') || str_contains($lowAddr, 'naroda') || str_contains($lowAddr, 'નરોડા')) {
                    if (!str_contains($lowAddr, 'new naroda') && !str_contains($lowAddr, 'nava naroda')) {
                        $resolvedAreaName = 'Naroda';
                    }
                }

                // Resolve Area ID
                $areaId = null;
                $pincode = null;

                if (!empty($resolvedAreaName)) {
                    $lookupName = strtolower($resolvedAreaName);
                    if (isset($areaMap[$lookupName])) {
                        $areaId = $areaMap[$lookupName]->id;
                        $pincode = $areaMap[$lookupName]->pincode;
                    } else {
                        $newArea = Area::create([
                            'name'    => ucwords($resolvedAreaName),
                            'pincode' => $canonicalPincodes[$lookupName] ?? null,
                        ]);
                        $areaMap[$lookupName] = $newArea;
                        $areaId = $newArea->id;
                        $pincode = $newArea->pincode;
                    }
                }

                // Extract pincode from address if empty
                $addressPincode = '380001';
                if (preg_match('/\b(38\d{4})\b/', $address, $pm)) {
                    $addressPincode = $pm[1];
                }
                if (empty($pincode)) {
                    $pincode = $addressPincode;
                }

                // Find user by member_code
                $user = User::where('member_code', $memberCode)->first();

                // Prevent email collision with existing users
                if ($email) {
                    $conflict = User::where('email', $email)->where(function($q) use ($user) {
                        if ($user) {
                            $q->where('id', '!=', $user->id);
                        }
                    })->exists();

                    if ($conflict) {
                        $email = null;
                    }
                }

                if (!$user) {
                    $user = User::create([
                        'name'           => $fullName,
                        'email'          => $email,
                        'password'       => $defaultPassword,
                        'status'         => 'approved',
                        'account_status' => 'open',
                        'member_code'    => $memberCode,
                    ]);
                    $user->assignRole($memberRole);
                } else {
                    $user->update([
                        'name'           => $fullName,
                        'email'          => $email,
                        'status'         => 'approved',
                        'account_status' => $user->account_status ?: 'open',
                        'member_code'    => $memberCode,
                    ]);
                    if (!$user->hasRole('Member') && !$user->hasRole('Administrator') && !$user->hasRole('Sub Admin')) {
                        $user->assignRole($memberRole);
                    }
                }

                // Update or create member profile
                MemberProfile::updateOrCreate(
                    ['user_id' => $user->id],
                    [
                        'first_name'  => $firstName ?: $fullName,
                        'middle_name' => $middleName,
                        'last_name'   => $surname,
                        'gender'      => 'Male',
                        'phone'       => $phone ?: 'N/A',
                        'whatsapp'    => $phone ?: 'N/A',
                        'address'     => $address ?: '',
                        'area_id'     => $areaId,
                        'pincode'     => $pincode ?: '380001',
                        'city'        => 'Ahmedabad',
                        'state'       => 'Gujarat',
                    ]
                );

                $importedUserIds[] = $user->id;
                $count++;
            }

            // Remove dummy members not in the dataset (keep Admin and Sub-Admin accounts safe)
            $dummyUsers = User::role('Member')
                ->whereDoesntHave('roles', function($q) {
                    $q->whereIn('name', ['Administrator', 'Sub Admin']);
                })
                ->whereNotIn('id', $importedUserIds)
                ->get();

            $deletedCount = 0;
            foreach ($dummyUsers as $du) {
                if ($du->memberProfile) {
                    $du->memberProfile->delete();
                }
                $du->delete();
                $deletedCount++;
            }

            DB::commit();

            if (isset($this->command)) {
                $this->command->info("Member import completed! Successfully seeded {$count} members. Cleaned duplicate areas.");
            }
        } catch (\Exception $e) {
            DB::rollBack();
            if (isset($this->command)) {
                $this->command->error("Failed to seed members: " . $e->getMessage());
            }
            throw $e;
        }
    }
}
"""

with open('database/seeders/MemberImportSeeder.php', 'w', encoding='utf-8') as f:
    f.write(php_template)

print("Generated database/seeders/MemberImportSeeder.php with automatic Area deduplication!")
